import os.path
import threading
from tkinter import ttk, filedialog
import tkinter as tk
from openpyxl import Workbook


# noinspection PyMethodMayBeStatic,PyShadowingNames
class TableConverterApp:
    """Utility to convert text table file to excel file
    """
    def __init__(self):
        """Initialize UI componants
        """
        self.root = root
        self.root.resizable(False, False)
        self.root.title("Table Converter")

        self.label_colour = "orange"
        self.input_label = tk.Label(root, text="Select a File:", fg=self.label_colour)
        self.input_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")

        self.entry_bg_color = "lightgray"
        entry_var = tk.StringVar()
        self.input_entry = ttk.Entry(root, textvariable=entry_var, width=40)
        self.input_entry.grid(row=0, column=1, padx=10, pady=5, sticky="we")

        self.output_label = tk.Label(root, text="Output File:", fg=self.label_colour)
        self.output_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")

        self.output_entry = tk.Entry(root, state=tk.DISABLED, width=40, bg=self.entry_bg_color)
        self.output_entry.grid(row=1, column=1, padx=10, pady=5, sticky="we")

        self.browse_button = tk.Button(root, text="Browse", command=self.browse_input_file, bg="gray", fg="white")
        self.browse_button.grid(row=0, column=2, padx=10, pady=5)

        self.convert_button = tk.Button(root, text="Convert", command=self.convert, bg="gray", fg="white")
        self.convert_button.grid(row=1, column=2, padx=10, pady=5)

        self.progress_label = tk.Label(root, text="Converting:", fg=self.label_colour)
        self.progress_label.grid(row=2, column=0, columnspan=3, padx=10, pady=5, sticky="w")

        self.progress_bar = ttk.Progressbar(root, orient="horizontal", length=300, mode="determinate", style="TProgressbar")
        self.s = ttk.Style()
        self.progress_bar.grid(row=3, column=0, columnspan=3, padx=10, pady=5)
        self.s.configure("TProgressbar", thickness=30, throughcolor="blue", background="green")

        self.progress_label_lock = threading.Lock()
        self.save_lock = threading.Lock()

    def update_progress(self, progress_label_text, progress_value):
        with self.progress_label_lock:
            self.progress_label.config(text=progress_label_text)
            self.progress_bar["value"] = progress_value
            self.root.update_idletasks()

    def browse_input_file(self):
        """File browse widget
        """
        file_path = filedialog.askopenfilename(filetypes=[("Text Files", "*.txt")])
        if file_path:
            self.input_entry.delete(0, tk.END)
            self.input_entry.insert(0, file_path)
            self.update_output_file_entry(file_path)

    def update_output_file_entry(self, input_file):
        """Set value/path to output file nme

        Args:
            input_file (str): input file name/path
        """
        input_path, input_filename = os.path.split(input_file)
        output_filename = os.path.splitext(input_filename)[0] + ".xlsx"
        self.output_entry.config(state=tk.NORMAL)
        self.output_entry.delete(0, tk.END)
        self.output_entry.insert(0, os.path.join(input_path, os.path.basename(output_filename)))
        self.output_entry.config(state=tk.DISABLED)
    
    def write_table_summary(self, tables, wb):
        summary_sheet = wb.active
        summary_sheet.title = "Summary"
        summary_sheet.append(["mismatch_column","mismatch_count"])

        for idx, table in enumerate(tables):
            header = table["header"]
            column_name = header[len(header) - 1].split('_onCloud')[0]
            mismatch_count = len(table["data"])
            summary_sheet.append([column_name, mismatch_count])

    def convert(self):
        """Start conversion to excel
        """
        input_file = self.input_entry.get()
        output_file = self.output_entry.get()
        max_rows_per_sheet = 250000

        tables = self.process_text_file(input_file)


        conversion_thread = threading.Thread(target=self.write_tables_to_excel,
                                             args=(tables, output_file, max_rows_per_sheet))
        conversion_thread.start()

    def write_tables_to_excel(self, tables, output_file, max_rows_per_sheet=100):
        """Write tables to excel file

        Args:
            tables (list): List of table data
            output_file (str): output file path
            max_rows_per_sheet (int, optional): maximum rows allowed per sheet to write. Defaults to 100.
        """
        wb = Workbook()
        table_name = ""
        total_tables = len(tables)
        table_number = 1
        threads = []

        self.write_table_summary(tables=tables, wb=wb)

        def write_table_thread(table, sheet_idx):
            """Thread to write tables to excel

            Args:
                table (dict): contains data of single table
                sheet_idx (int): sheet id
            """
            header = table['header']
            table_name = header[len(header) - 1].split('_onCloud')[0]
            data = table['data']
            num_rows = len(data)
            num_sheets = num_rows // max_rows_per_sheet + (num_rows % max_rows_per_sheet > 0)

            for sheet_idx in range(num_sheets):
                sheet_title = f'{table_name}_{sheet_idx + 1}'
                sheet = wb.create_sheet(title=sheet_title)
                start_row = sheet_idx * max_rows_per_sheet
                end_row = min(start_row + max_rows_per_sheet, num_rows)
                if start_row == 0 or start_row >= max_rows_per_sheet:
                    sheet.append(header)
                for row_data in data[start_row:end_row]:
                    sheet.append(row_data)

            progress_label_text = f"Converting: {table_name}"
            progress_value = (sheet_idx + 1) * 100/total_tables
            self.update_progress(progress_label_text=progress_label_text, progress_value=progress_value)


        for idx, table in enumerate(tables):
            thread = threading.Thread(target=write_table_thread, args=(table, idx))
            threads.append(thread)
            thread.start()

        for thread in threads:
            thread.join()

        if wb.sheetnames[0] == "Sheet":
            wb.remove(wb["Sheet"])

        with self.save_lock:
            wb.save(output_file)
        
        self.update_progress("Converting:",0)

    def process_text_file(self, input_file):
        """Convert text tabular file to dictionary of header and data

        Args:
            input_file (str): input file path

        Returns:
            dict: returns dictionary of header and data for each table
        """
        tables = []
        current_table = None

        with open(input_file, 'r') as f:
            for line in f:
                if line.startswith('|'):
                    row_data = self.get_row_data(line)
                    if current_table is None:
                        header = row_data
                        current_table = {'header': header, 'data': []}
                    else:
                        current_table['data'].append(row_data)
                elif line.startswith('[FAILED] Source and Target Data is not matching for column'):
                    if current_table:
                        tables.append(current_table)
                    current_table = None

        if current_table:
            tables.append(current_table)

        return tables
    
    def get_row_data(self, line):
        """Convert row in text file to row data

        Args:
            line (str): line from table(row)

        Returns:
            list: list(tuple) of row data
        """
        row_data = []
        splitted_line = [colmn.strip() for colmn in line.strip().split(' | ')]
        for item in splitted_line:
            if item.startswith("| "):
                item_split = item.split("| ")
                row_data.append(item_split[1].strip())
            elif item.endswith("|"):
                item_split = item.split(" |")
                row_data.append(item_split[0].strip())
            else:
                row_data.append(item)
        return row_data


if __name__ == "__main__":
    root = tk.Tk()
    app = TableConverterApp()
    root.mainloop()